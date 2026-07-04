"""从 Excel 导入作者文章到 author_articles 表"""

import sys
from pathlib import Path

# 添加 backend 到 path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from db import create_author_article, get_author_article_by_url, list_author_articles


EXCEL_PATH = Path(__file__).parent.parent / "微信公众号文章纯文字.xlsx"


def import_from_excel(excel_path: str = None) -> dict:
    """读取 Excel 并导入到 author_articles 表。

    返回: {"total": int, "imported": int, "skipped": int}
    """
    path = Path(excel_path) if excel_path else EXCEL_PATH
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    total = 0
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        url = (data.get("链接") or "").strip()
        if not url:
            continue

        total += 1

        # 去重检查
        existing = get_author_article_by_url(url)
        if existing:
            skipped += 1
            continue

        # 解析发布时间
        publish_time = data.get("发布时间") or ""
        if publish_time and not isinstance(publish_time, str):
            publish_time = str(publish_time)

        # 解析数值
        read_count = data.get("阅读")
        like_count = data.get("点赞")
        if read_count is not None:
            read_count = int(read_count)
        if like_count is not None:
            like_count = int(like_count)

        create_author_article(
            url=url,
            title=(data.get("标题") or "").strip(),
            publish_time=publish_time,
            summary=(data.get("摘要") or "").strip(),
            article_type=(data.get("文章类型") or "").strip(),
            tags=(data.get("所属合集") or "").strip(),
            read_count=read_count,
            like_count=like_count,
        )
        imported += 1

    wb.close()
    return {"total": total, "imported": imported, "skipped": skipped}


if __name__ == "__main__":
    result = import_from_excel()
    print(f"导入完成: 总计 {result['total']} 篇, 新增 {result['imported']} 篇, 跳过 {result['skipped']} 篇")
